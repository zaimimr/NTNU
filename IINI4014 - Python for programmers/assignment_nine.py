from openpyxl import load_workbook
wb = load_workbook(filename='Fertility .xlsx', read_only=True)
ws = wb['Fertility ']

class Country:
    def __init__(self, name, fertility):
        self.name = name
        self.fertility = fertility

cName = []
cFertility = []
mRows = ws.max_row
#Loop to find the conutries for the excel sheet and also find total fereility. Will be moved to two seperate lists.
for cells in range(2, mRows):
    ca = ws['A' + str(cells+1)].value
    cb = ws['A' + str(cells+2)].value
    if ca is not cb:
        tFertility = ws.cell(row=cells+1, column = 6).value
        cName.append(ca)
        cFertility.append(tFertility)
    else:
        continue


land = []
#Loop to make Country objects and moved to a list.
for i in range(len(cName)):
    land.append(Country(cName[i], cFertility[i]))


#Highest fertility
for x in range(len(cName)):
    if land[x].fertility == max(cFertility):
        print('Highest fertility: ', land[x].name, land[x].fertility)
    else:
        continue

#Lowest fertility
for y in range(len(cName)):
    if land[y].fertility == min(cFertility):
        print('Lowest fertility: ', land[y].name, land[y].fertility)
    else:
        continue
